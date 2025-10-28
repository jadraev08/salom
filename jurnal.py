


import tkinter as tk
from tkinter import ttk, messagebox
import openpyxl
import os

FILENAME = "jurnal.xlsx"

class JurnalApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Jurnal Ilovasi")
        self.root.geometry("1200x500")

        columns = ["Ism", "Telefon"] + [f"Dars {i}" for i in range(1, 13)] + ["Test natijasi"]
        self.tree = ttk.Treeview(root, columns=columns, show="headings")

        for col in columns:
            self.tree.heading(col, text=col)
            self.tree.column(col, width=80, anchor="center")

        self.tree.pack(fill="both", expand=True)

        form_frame = tk.Frame(root)
        form_frame.pack(pady=10)

        tk.Label(form_frame, text="Ism:").grid(row=0, column=0)
        self.name_entry = tk.Entry(form_frame)
        self.name_entry.grid(row=0, column=1)

        tk.Label(form_frame, text="Telefon:").grid(row=0, column=2)
        self.phone_entry = tk.Entry(form_frame)
        self.phone_entry.grid(row=0, column=3)

        tk.Label(form_frame, text="Test natijasi:").grid(row=0, column=4)
        self.test_entry = tk.Entry(form_frame)
        self.test_entry.grid(row=0, column=5)

        add_btn = tk.Button(form_frame, text="Qo'shish", command=self.add_student)
        add_btn.grid(row=0, column=6, padx=10)

        mark_present_btn = tk.Button(form_frame, text="Keldi ✅", command=lambda: self.mark_attendance("✅"))
        mark_present_btn.grid(row=0, column=7, padx=5)

        mark_absent_btn = tk.Button(form_frame, text="Kelmadi ❌", command=lambda: self.mark_attendance("❌"))
        mark_absent_btn.grid(row=0, column=8, padx=5)

        save_btn = tk.Button(root, text="Saqlash (Excel)", command=self.save_to_excel)
        save_btn.pack(pady=5)

        # Agar fayl mavjud bo‘lsa, o‘qib yuklash
        if os.path.exists(FILENAME):
            self.load_from_excel()

    def add_student(self):
        name = self.name_entry.get()
        phone = self.phone_entry.get()
        test = self.test_entry.get()

        if not name or not phone:
            messagebox.showwarning("Xato", "Ism va telefon raqamini kiriting!")
            return

        row = [name, phone] + ["" for _ in range(12)] + [test]
        self.tree.insert("", "end", values=row)

        self.name_entry.delete(0, tk.END)
        self.phone_entry.delete(0, tk.END)
        self.test_entry.delete(0, tk.END)

    def mark_attendance(self, status):
        selected = self.tree.selection()
        if not selected:
            messagebox.showwarning("Xato", "O'quvchini tanlang!")
            return

        item = selected[0]
        values = list(self.tree.item(item, "values"))

        for i in range(2, 14):
            if values[i] == "":
                values[i] = status
                break

        self.tree.item(item, values=values)

    def save_to_excel(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Jurnal"

        # Ustun sarlavhalar
        columns = ["Ism", "Telefon"] + [f"Dars {i}" for i in range(1, 13)] + ["Test natijasi"]
        ws.append(columns)

        # Treeviewdagi barcha ma’lumotlarni yozish
        for row in self.tree.get_children():
            ws.append(self.tree.item(row)["values"])

        wb.save(FILENAME)
        messagebox.showinfo("Muvaffaqiyatli", f"Ma'lumotlar {FILENAME} fayliga saqlandi!")

    def load_from_excel(self):
        wb = openpyxl.load_workbook(FILENAME)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            self.tree.insert("", "end", values=row)

if __name__ == "__main__":
    root = tk.Tk()
    app = JurnalApp(root)
    root.mainloop()



